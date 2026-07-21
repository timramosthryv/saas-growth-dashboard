#!/usr/bin/env python3
"""
build.py - SaaS Growth Dashboard builder

Usage: python build.py data.xlsx
Output: index.html (self-contained; committed to GitHub Pages)

Reads the SaaS Growth export, then:
  1. keeps only the Reviewed and Reported 2026 sections,
  2. drops any growth that OCCURRED before 2026-01-01 (the pre-2026 floor),
  3. derives the dashboard columns (SaaS Growth, PCSM, Coach, Month Impacted, Quarter),
  4. injects the data into template.html.

The transform is idempotent, so it works whether the input file is a raw export
or one that already contains the derived columns.
"""
import sys, os, json
from datetime import datetime, date

KEEP_SECTIONS = {"Reviewed", "Reported 2026"}
FLOOR = date(2026, 1, 1)   # drop growths that occurred before this date
MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def serialize(v):
    if isinstance(v, (datetime, date)):
        return str(v)
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def to_num(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_date(v):
    if not v:
        return None
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def load_rows(path):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb['Raw Data'] if 'Raw Data' in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = {h: r[i] for i, h in enumerate(headers) if h}
        if any(v is not None and v != "" for v in row.values()):
            rows.append(row)
    return rows


def transform(rows):
    out = []
    for row in rows:
        if row.get('Section') not in KEEP_SECTIONS:
            continue
        occ = parse_date(row.get('Date SaaS Growth Occured'))
        if occ is not None and occ < FLOOR:
            continue  # pre-2026 floor
        new = to_num(row.get('New SaaS Dollars')) or 0
        old = to_num(row.get('Old SaaS Dollars')) or 0
        row['SaaS Growth'] = new - old
        row['PCSM'] = row.get('PCSM Submitter') or ''
        row['Coach'] = row.get('Assignee') or ''
        if occ is not None:
            row['Month Impacted'] = f"{MONTHS[occ.month - 1]} {occ.year}"
            row['Quarter'] = f"Q{(occ.month - 1) // 3 + 1} {occ.year}"
        else:
            row['Month Impacted'] = ''
            row['Quarter'] = ''
        obj = {h: serialize(v) for h, v in row.items() if v is not None and v != ""}
        if obj:
            out.append(obj)
    return out


def build(xlsx_path):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, 'template.html')
    output_path = os.path.join(script_dir, 'index.html')
    if not os.path.exists(template_path):
        raise FileNotFoundError("template.html not found next to build.py")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"{xlsx_path} not found")
    print(f"Loading {xlsx_path} ...")
    rows = transform(load_rows(xlsx_path))
    print(f"  {len(rows)} rows after filter + floor + transform")
    with open(template_path, 'r', encoding='utf-8') as f:
        html = f.read()
    ts = datetime.now().strftime('%b %d, %Y')
    json_data = json.dumps(rows, separators=(',', ':'), ensure_ascii=False)
    injection = f"// Embedded dataset - generated {ts}\nconst __EMBEDDED_DATA__={json_data};"
    html = html.replace('// @@DATA_INJECTION@@', injection)
    html = html.replace("DATA_TS='@@TIMESTAMP@@'", f"DATA_TS='{ts}'")
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  index.html written ({os.path.getsize(output_path) // 1024} KB)")
    print("Done!")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python build.py data.xlsx")
        sys.exit(1)
    build(sys.argv[1])
